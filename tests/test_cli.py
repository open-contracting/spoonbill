import logging
import os
import pathlib
import shutil
from unittest import mock

import openpyxl
from click.testing import CliRunner

from spoonbill.cli import cli
from spoonbill.utils import RepeatFilter

LOGGER = logging.getLogger("spoonbill")
LOGGER.addFilter(RepeatFilter())

FILENAME = pathlib.Path("tests/data/ocds-sample-data.json").absolute()
FILENAME_JSONL = pathlib.Path("tests/data/ocds-sample-data.jsonl").absolute()
FILENAME_GZ = pathlib.Path("tests/data/ocds-sample-data.json.gz").absolute()
EMPTY_LIST_FILE = pathlib.Path("tests/data/empty_list.json").absolute()
SCHEMA = pathlib.Path("tests/data/ocds-simplified-schema.json").absolute()
ANALYZED = pathlib.Path("tests/data/analyzed.state").absolute()
ONLY = pathlib.Path("tests/data/only.txt").absolute()
UNNEST = pathlib.Path("tests/data/unnest.txt").absolute()


def test_no_filename():
    runner = CliRunner()
    result = runner.invoke(cli, [])
    assert result.exit_code == 2
    assert result.output.startswith("Usage")
    assert "Missing argument 'FILENAME'" in result.output


def test_default():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Dumping analyzed data" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output


def test_with_selections():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["--selection", "tenders", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Dumping analyzed data" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output


def test_with_exclude():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["--exclude", "tenders_items", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Ignoring tables (excluded by user): tenders_items" in result.output
        assert "Dumping analyzed data" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output


def test_with_combine():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["--combine", "documents", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Dumping analyzed data" in result.output
        assert (
            "Going to export tables: parties,parties_ids,planning,tenders,awards,contracts,documents" in result.output
        )
        assert "Done flattening. Flattened objects: 6" in result.output


def test_table_typo():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        for option in ("--selection", "--combine"):
            result = runner.invoke(cli, [option, "missing", "data.json"])
            assert result.exit_code == 2
            assert "Invalid value" in result.output
            assert "Wrong selection, table 'missing' does not exist"


def test_with_schema():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(SCHEMA, "schema.json")
        result = runner.invoke(cli, ["--selection", "tenders", "--schema", "schema.json", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Dumping analyzed data" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output
        assert "Going to export tables: tenders" in result.output


def test_state_file():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        result = runner.invoke(cli, ["--state-file", "analyzed.json", "data.json"])
        assert result.exit_code == 0
        assert "Restoring from provided state file" in result.output
        assert (
            "Going to export tables: parties,planning,tenders,awards,contracts,documents,milestones,amendments"
            in result.output
        )
        assert "Done flattening. Flattened objects: 6" in result.output


def test_only():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        result = runner.invoke(cli, ["--only", "/tender/title", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Using only columns" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output


def test_only_file():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        shutil.copyfile(ONLY, "only")
        result = runner.invoke(cli, ["--only-file", "only", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Using only columns" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output


def test_repleat():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        result = runner.invoke(cli, ["--repeat", "/tender/id", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Repeating columns /tender/id in all child table of tenders" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output


def test_repeat_file():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        shutil.copyfile(ONLY, "only")
        result = runner.invoke(cli, ["--repeat-file", "only", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Repeating columns /tender/id,/tender/title in all child table of tenders" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output


def test_unnest():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        result = runner.invoke(cli, ["--unnest", "/tender/items/0/id", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Unnesting columns /tender/items/0/id for table tenders" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output


def test_unnest_file():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        shutil.copyfile(UNNEST, "unnest")
        result = runner.invoke(cli, ["--unnest-file", "unnest", "data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output
        assert "Unnesting columns /tender/items/0/id for table tenders" in result.output
        assert "Done flattening. Flattened objects: 6" in result.output


def test_table_stats():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["data.json"])
        for msg in (
            "Analyze options:",
            " - threshold                      => 5",
            " - language                       => en",
        ):
            assert msg in result.output
        for msg in (
            "Processed tables:",
            " - tenders                        => 4 rows",
            " - awards                         => 3 rows",
            " - contracts                      => 2 rows",
            " - planning                       => 1 rows",
            " - parties                        => 8 rows",
            " ---- parties_ids                 => 14 rows",
        ):
            assert msg in result.output


def test_message_repeat(capsys):
    message = "Around the world, around the world"

    LOGGER.warning(message)
    LOGGER.warning(message)
    LOGGER.warning(message)
    captured = capsys.readouterr()
    assert captured.out.count(message) == 1


def test_xlsx():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        result = runner.invoke(cli, ["--state-file", "analyzed.json", "--xlsx", "test.xlsx", "data.json"])
        assert result.exit_code == 0
        assert "Done flattening. Flattened objects: 6" in result.output
        path = pathlib.Path("test.xlsx")
        assert path.resolve().exists()


def test_csv():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        os.mkdir("test")
        result = runner.invoke(cli, ["--state-file", "analyzed.json", "--csv", "test", "data.json"])
        assert result.exit_code == 0
        assert "Done flattening. Flattened objects: 6" in result.output
        path = pathlib.Path("test").resolve() / "tenders.csv"
        assert path.resolve().exists()


def test_empty_list_file():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(EMPTY_LIST_FILE, "data.json")
        result = runner.invoke(cli, ["data.json"])
        assert result.exit_code == 0


def test_default_field_only():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["--selection", "tenders,parties", "--only", "parentID", "data.json"])
        assert "Using only columns parentID for table tenders" in result.output
        assert "Using only columns parentID for table parties" in result.output


def test_jsonl():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME_JSONL, "data.json")
        result = runner.invoke(cli, ["data.json"])
        assert "Input file is release" in result.output
        assert result.exit_code == 0


def test_gz():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME_GZ, "data.json.gz")
        result = runner.invoke(cli, ["data.json.gz"])
        assert result.exit_code == 0


def test_sheet_order():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["data.json"])
        assert result.exit_code == 0
        path = pathlib.Path("result.xlsx")
        assert path.resolve().exists()
        workbook = openpyxl.load_workbook(path)
        sheets = [name for name in workbook.sheetnames]
        assert sheets == [
            "parties",
            "parties_ids",
            "planning",
            "tenders",
            "awards",
            "contracts",
            "documents",
            "milestones",
            "amendments",
        ]


def test_sheet_order_state_file():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        shutil.copyfile(ANALYZED, "analyzed.json")
        result = runner.invoke(cli, ["--state-file", "analyzed.json", "data.json"])
        assert result.exit_code == 0
        path = pathlib.Path("result.xlsx")
        assert path.resolve().exists()
        workbook = openpyxl.load_workbook(path)
        sheets = [name for name in workbook.sheetnames]
        assert sheets == [
            "parties",
            "planning",
            "tenders",
            "awards",
            "contracts",
            "documents",
            "milestones",
            "amendments",
        ]


@mock.patch.dict(os.environ, {"LANG": "uk_UA"})
def test_locale_not_found():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["data.json"])
        assert result.exit_code == 0
        assert "Input file is release package" in result.output


def test_pretty_headers_en():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["--human", "--language", "en", "data.json"])
        assert result.exit_code == 0
        path = pathlib.Path("result.xlsx")
        assert path.resolve().exists()
        xlsx_reader = openpyxl.load_workbook(path)
        sheet = xlsx_reader["parties"]
        header_values = [cell.value for cell in sheet[1]]
        header_columns = [cell.column_letter for cell in sheet[1]]
        headers = dict(zip(header_columns, header_values))
        assert headers["E"].startswith("Buyer")


def test_pretty_headers_es():
    runner = CliRunner()
    with runner.isolated_filesystem():
        shutil.copyfile(FILENAME, "data.json")
        result = runner.invoke(cli, ["--human", "--language", "es", "data.json"])
        assert result.exit_code == 0
        path = pathlib.Path("result.xlsx")
        assert path.resolve().exists()
        xlsx_reader = openpyxl.load_workbook(path)
        sheet = xlsx_reader["parties"]
        header_values = [cell.value for cell in sheet[1]]
        header_columns = [cell.column_letter for cell in sheet[1]]
        headers = dict(zip(header_columns, header_values))
        assert headers["E"].startswith("Comprador")
