import csv
from pathlib import Path
from unittest.mock import call, patch

import openpyxl
from scalpl import Cut

from spoonbill import FileAnalyzer, FileFlattener
from spoonbill.flatten import Flattener, FlattenOptions
from spoonbill.utils import SchemaHeaderExtractor, add_paths_to_schema, generate_paths
from spoonbill.writers.csv import CSVWriter
from spoonbill.writers.xlsx import XlsxWriter

from .conftest import releases_extension_path, releases_path
from .utils import get_writers, prepare_tables, read_csv_headers, read_xlsx_headers

ID_FIELDS = {"tenders": "/tender/id", "parties": "/parties/id"}


def test_writer_init(spec, tmpdir, flatten_options):
    tables = prepare_tables(spec, flatten_options, ID_FIELDS)
    workdir = Path(tmpdir)
    get_writers(workdir, tables, flatten_options)
    path = workdir / "result.xlsx"

    assert path.is_file()
    for name in flatten_options.selection:
        path = workdir / f"{name}.csv"
        assert path.is_file()


def test_headers_filtering(spec, tmpdir, flatten_options):
    tables = prepare_tables(spec, flatten_options, ID_FIELDS)
    workdir = Path(tmpdir)
    get_writers(workdir, tables, flatten_options)
    for name in flatten_options.selection:
        path = workdir / f"{name}.csv"
        xlsx_headers = read_xlsx_headers(workdir / "result.xlsx", name)
        csv_headers = read_csv_headers(path)
        assert len(xlsx_headers) == 1
        assert len(csv_headers) == 1
        assert xlsx_headers[0] == ID_FIELDS[name]
        assert csv_headers[0] == ID_FIELDS[name]


def test_writers_pretty_headers(spec, tmpdir, releases, schema):
    # increase items count for force split
    releases[0]["tender"]["items"] = releases[0]["tender"]["items"] * 6
    for _ in spec.process_items(releases):
        pass
    options = FlattenOptions(
        **{
            "selection": {
                "tenders": {"split": True, "pretty_headers": True},
                "parties": {"split": False, "pretty_headers": True},
                "tenders_items": {
                    "split": True,
                    "headers": {"/tender/items/id": "item id"},
                    "pretty_headers": True,
                },
            }
        }
    )
    tables = {
        "tenders": spec.tables["tenders"],
        "parties": spec.tables["parties"],
        "tenders_items": spec.tables["tenders_items"],
    }
    tables_headers = {
        "tenders": [
            "Ocid",
            "Id",
            "Row Id",
            "Parent Id",
            "Tender: Award Criteria",
            "Tender: Award Criteria Details",
            "Tender: Tender Description",
            "Tender: Has Enquiries?",
            "Tender: Tender Id",
            "Tender: Main Procurement Category",
            "Tender: Number Of Tenderers",
            "Tender: Procurement Method",
            "Tender: Procurement Method Details",
            "Tender: Procurement Method Rationale",
            "Tender: Tender Status",
            "Tender: Submission Method",
            "Tender: Submission Method Details",
            "Tender: Tender Title",
            "Tender: Value: Amount",
            "Tender: Value: Currency",
            "Tender: Tender Period: Duration ( Days)",
            "Tender: Tender Period: End Date",
            "Tender: Tender Period: Start Date",
            "Tender: Tenderers: Organization Id",
            "Tender: Tenderers: Organization Name",
            "Tender: Value: Amount",
            "Tender: Value: Currency",
            "Tender: Tender Period: Duration ( Days)",
            "Tender: Tender Period: End Date",
            "Tender: Tender Period: Start Date",
            "Tender: Tender Period: Duration ( Days)",
            "Tender: Tender Period: End Date",
            "Tender: Tender Period: Start Date",
            "Tender: Tender Period: End Date",
            "Tender: Tender period: Start date",
        ],
        "parties": [
            "Ocid",
            "Id",
            "Row Id",
            "Parent Id",
            "Parties: Organization: Entity Id",
            "Parties: Organization: Common Name",
            "Parties: Organization: Party Roles",
            "Parties: Organization: Primary Identifier: Id",
            "Parties: Organization: Primary Identifier: Legal Name",
            "Parties: Organization: Primary Identifier: Scheme",
            "Parties: Organization: Contact Point: Email",
            "Parties: Organization: Contact Point: Fax Number",
            "Parties: Organization: Contact Point: Name",
            "Parties: Organization: Contact Point: Telephone",
            "Parties: Organization: Contact Point: Url",
            "Parties: Organization: Address: Country Name",
            "Parties: Organization: Address: Locality",
            "Parties: Organization: Address: Postal Code",
            "Parties: Organization: Address: Region",
            "Parties: Organization: Address: Street Address",
            "Buyer: Organization Id",
            "Buyer: Organization Name",
            "Parties: Test",
        ],
        "tenders_items": [
            "Ocid",
            "Id",
            "Row Id",
            "Parent Id",
            "Parent Table",
            "Tender: Items To Be Procured: Item: Description",
            "item id",
            "Tender: Items To Be Procured: Item: Quantity",
            "Tender: Items To Be Procured: Item: Unit: Id",
            "Tender: Items To Be Procured: Item: Unit: Name",
            "Tender: Items To Be Procured: Item: Unit: Scheme",
            "Tender: Value: Amount",
            "Tender: Value: Currency",
        ],
    }

    workdir = Path(tmpdir)
    get_writers(workdir, tables, options, schema)
    xlsx = workdir / "result.xlsx"

    for name, opts in options.selection.items():
        path = workdir / f"{name}.csv"
        xlsx_headers = read_xlsx_headers(xlsx, name)
        csv_headers = read_csv_headers(path)
        headers = tables_headers[name]
        assert not set(headers).difference(set(xlsx_headers))
        assert not set(headers).difference(set(csv_headers))

    options = FlattenOptions(
        **{
            "selection": {
                "tenders": {
                    "split": True,
                    "headers": {"/tender/id": "TENDER"},
                    "pretty_headers": True,
                },
                "tenders_items": {
                    "split": True,
                    "headers": {"/tender/items/id": "item id"},
                    "pretty_headers": True,
                },
                "parties": {
                    "split": False,
                    "headers": {"/parties/id": "PARTY"},
                    "pretty_headers": True,
                },
            }
        }
    )

    workdir = Path(tmpdir)
    get_writers(workdir, tables, options, schema)
    xlsx = workdir / "result.xlsx"

    sheet = "tenders"
    path = workdir / f"{sheet}.csv"
    xlsx_headers = read_xlsx_headers(xlsx, sheet)
    csv_headers = read_csv_headers(path)
    assert "TENDER" in xlsx_headers
    assert "TENDER" in csv_headers

    sheet = "tenders_items"
    path = workdir / f"{sheet}.csv"
    xlsx_headers = read_xlsx_headers(xlsx, sheet)
    csv_headers = read_csv_headers(path)
    assert "item id" in xlsx_headers
    assert "item id" in csv_headers

    xlsx = workdir / "result.xlsx"
    sheet = "parties"
    path = workdir / f"{sheet}.csv"

    xlsx_headers = read_xlsx_headers(xlsx, sheet)
    csv_headers = read_csv_headers(path)

    assert "PARTY" in xlsx_headers
    assert "PARTY" in csv_headers


def test_writers_flatten_count(spec, tmpdir, releases, schema):
    releases[0]["tender"]["items"] = releases[0]["tender"]["items"] * 6
    for _ in spec.process_items(releases):
        pass
    options = FlattenOptions(
        **{
            "selection": {
                "tenders": {"split": True, "pretty_headers": True},
                "parties": {"split": True, "pretty_headers": True},
                "tenders_items": {
                    "split": False,
                    "pretty_headers": True,
                },
                "parties_ids": {
                    "split": False,
                    "pretty_headers": True,
                },
            },
            "count": True,
        }
    )

    workdir = Path(tmpdir)
    analyzer = FileAnalyzer(workdir)
    flattener = FileFlattener(
        workdir=workdir, options=options, tables=spec.tables, csv=True, analyzer=analyzer, schema=schema
    )
    xlsx = workdir / "result.xlsx"
    for _ in flattener.flatten_file(releases_path):
        pass
    sheet = "tenders"
    path = workdir / f"{sheet}.csv"
    # TODO: xlsx headers
    headers = read_csv_headers(path)
    assert "Tender: Items Count" in headers
    assert "Tender: Tenderers Count" in headers

    sheet = "parties"
    path = workdir / f"{sheet}.csv"
    for headers in read_xlsx_headers(xlsx, sheet), read_csv_headers(path):
        assert "Parties: Additional Identifiers Count" in headers


def test_writers_table_name_override(spec, tmpdir, schema):
    options = FlattenOptions(
        **{
            "selection": {
                "parties": {"split": False, "pretty_headers": True, "name": "testname"},
                "tenders": {"split": True, "pretty_headers": True, "name": "my_tenders"},
                "tenders_items": {"split": False, "pretty_headers": True, "name": "pretty_items"},
            }
        }
    )
    tables = prepare_tables(spec, options)
    for name, table in tables.items():
        for col in table:
            table.inc_column(col, col)

    workdir = Path(tmpdir)
    get_writers(workdir, tables, options, schema)
    xlsx = workdir / "result.xlsx"
    for name in ("testname", "my_tenders", "pretty_items"):
        path = workdir / f"{name}.csv"
        assert path.is_file()
        assert read_xlsx_headers(xlsx, name)
        assert read_csv_headers(path)


def test_abbreviations(spec, tmpdir):
    options = FlattenOptions(
        **{
            "selection": {
                "tenders_items_class": {"split": False},
                "parties_ids": {"split": False},
                "transactions": {"split": False},
            }
        }
    )
    new_names = ["tenders_items_class", "parties_ids", "transactions"]
    tables = prepare_tables(spec, options)
    for name, table in tables.items():
        for col in table:
            table.inc_column(col, col)

    workdir = Path(tmpdir)
    get_writers(workdir, tables, options)
    xlsx = workdir / "result.xlsx"
    for name in new_names:
        path = workdir / f"{name}.csv"
        assert path.is_file()
        assert read_xlsx_headers(xlsx, name)
        assert read_csv_headers(path)


def test_name_duplicate(spec, tmpdir, schema):
    duplicate_name = "test"
    options = FlattenOptions(
        **{
            "selection": {
                "parties": {"split": False, "pretty_headers": True, "name": duplicate_name},
                "tenders": {"split": True, "pretty_headers": True, "name": duplicate_name},
                "tenders_items": {"split": False, "pretty_headers": True, "name": duplicate_name},
            }
        }
    )
    tables = prepare_tables(spec, options)
    for name, table in tables.items():
        for col in table:
            table.inc_column(col, col)
    workdir = Path(tmpdir)
    get_writers(workdir, tables, options, schema)
    xlsx = workdir / "result.xlsx"
    for name in ("test", "test1", "test2"):
        path = workdir / f"{name}.csv"
        assert path.is_file()
        assert read_xlsx_headers(xlsx, name)
        assert read_csv_headers(path)


@patch("spoonbill.LOGGER.error")
def test_writers_invalid_table(log, spec, tmpdir, schema):
    options = FlattenOptions(
        **{
            "selection": {
                "parties": {"split": False, "pretty_headers": True, "name": "testname"},
            }
        }
    )
    tables = prepare_tables(spec, options)
    for name, table in tables.items():
        for col in table:
            table.inc_column(col, col)

    workdir = Path(tmpdir)
    writers = get_writers(workdir, tables, options, schema)
    for writer in writers:
        writer.writerow("test", {})
    log.assert_has_calls([call("Invalid table test"), call("Invalid table test")])


@patch("spoonbill.LOGGER.error")
def test_writers_invalid_row(log, spec, tmpdir, schema):
    options = FlattenOptions(
        **{
            "selection": {
                "parties": {"split": False, "pretty_headers": True, "name": "testname"},
            }
        }
    )
    tables = prepare_tables(spec, options)
    for name, table in tables.items():
        for col in table:
            table.inc_column(col, col)

    workdir = Path(tmpdir)
    writers = get_writers(workdir, tables, options, schema)
    for writer in writers:
        writer.writerow("parties", {"/test/test": "test"})
    log.assert_has_calls(
        [
            call("Operation produced invalid path. This a software bug, please send issue to developers"),
            call("Failed to write row None with error dict contains fields not in fieldnames: '/test/test'"),
            call("Operation produced invalid path. This a software bug, please send issue to developers"),
            call("Failed to write column /test/test to xlsx sheet parties"),
        ]
    )


@patch("spoonbill.LOGGER.error")
@patch("spoonbill.writers.csv.open", **{"side_effect": OSError("test")})
def test_writers_open_fail(open_, log, spec, tmpdir):
    options = FlattenOptions(
        **{
            "selection": {
                "parties": {"split": False, "pretty_headers": True, "name": "testname"},
            }
        }
    )
    workdir = Path(tmpdir)
    tables = prepare_tables(spec, options)
    get_writers(workdir, tables, options)
    log.assert_has_calls([call("Failed to open file {} with error {}".format(str(tmpdir / "testname.csv"), "test"))])


def test_csv_writer(spec_analyzed, releases, flatten_options, tmpdir, schema):
    flattener = Flattener(flatten_options, spec_analyzed.tables)
    flatten_options.selection["parties"].split = True
    tables = prepare_tables(spec_analyzed, flatten_options)
    workdir = Path(tmpdir)
    with CSVWriter(workdir, tables, flatten_options, schema) as writer:
        # Writing CSV files
        for _count, flat in flattener.flatten(releases):
            for name, rows in flat.items():
                for row in rows:
                    writer.writerow(name, row)

    # Reading CSV files
    counter = {}
    for _count, flat in flattener.flatten(releases):
        for name, rows in flat.items():
            if name not in counter:
                counter[name] = 0
            for row in rows:
                str_row = {k: str(v) for (k, v) in row.items()}
                file = name + ".csv"
                path = workdir / file
                with open(path, newline="") as csv_file:
                    csv_reader = csv.DictReader(csv_file)
                    for num, line in enumerate(csv_reader):
                        if num == counter[name]:
                            clean_line = {k: v for (k, v) in line.items() if v != ""}
                            assert dict(clean_line) == str_row
                counter[name] += 1


def test_xlsx_writer(spec_analyzed, releases, flatten_options, tmpdir, schema):
    flattener = Flattener(flatten_options, spec_analyzed.tables)
    tables = prepare_tables(spec_analyzed, flatten_options)
    workdir = Path(tmpdir)
    with XlsxWriter(workdir, tables, flatten_options, schema) as writer:
        # Writing XLSX file
        for _count, flat in flattener.flatten(releases):
            for name, rows in flat.items():
                for row in rows:
                    writer.writerow(name, row)

    # Reading XLSX files
    counter = {}
    path = workdir / "result.xlsx"
    xlsx_reader = openpyxl.load_workbook(path)

    for _count, flat in flattener.flatten(releases):
        for name, rows in flat.items():
            if name not in counter:
                counter[name] = 2
            sheet = xlsx_reader[name]
            headers = {cell.column_letter: cell.value for cell in sheet[1]}
            for row in rows:
                line = {headers[cell.column_letter]: cell.value for cell in sheet[counter[name]]}
                row = row.as_dict()
                assert not set(row.keys()).difference(set(line.keys()))
                for k, v in row.items():
                    assert str(v) == str(line[k])
                counter[name] += 1


def test_xlsx_only_no_default_columns(spec_analyzed, releases, tmpdir, schema):
    flatten_options = FlattenOptions(**{"selection": {"tenders": {"split": True, "only": ["/tender/id"]}}})
    flattener = Flattener(flatten_options, spec_analyzed.tables)
    tables = prepare_tables(spec_analyzed, flatten_options)
    workdir = Path(tmpdir)
    with XlsxWriter(workdir, tables, flatten_options, schema) as writer:
        for _count, flat in flattener.flatten(releases):
            for name, rows in flat.items():
                for row in rows:
                    writer.writerow(name, row)

    path = workdir / "result.xlsx"
    xlsx_reader = openpyxl.load_workbook(path)
    column = []
    for row in xlsx_reader["tenders"].rows:
        column.append(row[0].value)
    assert column[0] == "/tender/id"
    assert xlsx_reader["tenders"].max_column == 1


def test_flatten_multiple_files(spec, tmpdir, releases, schema):
    for _ in spec.process_items(releases):
        pass
    options = FlattenOptions(**{"selection": {"tenders": {"split": True}}})

    workdir = Path(tmpdir)
    analyzer = FileAnalyzer(workdir)
    flattener = FileFlattener(
        workdir=workdir, options=options, tables=spec.tables, csv=True, analyzer=analyzer, schema=schema
    )
    xlsx = workdir / "result.xlsx"
    sheet = "tenders"
    for _ in flattener.flatten_file(releases_path):
        pass
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    line_number = ws.max_row - 1
    assert ws.max_row - 1 == 4

    flattener = FileFlattener(
        workdir=workdir, options=options, tables=spec.tables, csv=True, analyzer=analyzer, schema=schema
    )
    for _ in flattener.flatten_file([releases_path, releases_path]):
        pass
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    assert ws.max_row - 1 == line_number * 2


def test_extension_export(spec, tmpdir, releases_extension, schema):
    for _ in spec.process_items(releases_extension):
        pass
    options = FlattenOptions(**{"selection": {"tenders": {"split": False}, "documents": {"split": False}}})

    workdir = Path(tmpdir)
    analyzer = FileAnalyzer(workdir)
    flattener = FileFlattener(workdir=workdir, options=options, tables=spec.tables, analyzer=analyzer, schema=schema)
    xlsx = workdir / "result.xlsx"
    sheet = "documents"
    extension_header = "/documents/test_extension"
    for _ in flattener.flatten_file(releases_extension_path):
        pass
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    for column_cell in ws.iter_cols(1, ws.max_column):
        if column_cell[0].value == extension_header:
            extension_column = ws[column_cell[0].coordinate[:-1]]
    for cell in extension_column:
        if cell.value == extension_header:
            continue
        assert cell.value == "test"
    for column_cell in wb["tenders"].iter_cols(1, ws.max_column):
        assert column_cell[0].value != extension_header


def test_schema_header_paths(schema):
    paths = generate_paths(schema["properties"])
    schema = add_paths_to_schema(schema)
    proxy = Cut(schema["properties"])
    for path in paths:
        if path[-1] == "title":
            _path = ".".join(path[:-1])
            assert "$title" in proxy[_path]
            assert proxy[_path]["$title"][-1] == path


def test_schema_header_generation(schema):
    headers = SchemaHeaderExtractor(schema)
    paths = [
        ["parties", "title"],
        ["parties", "items", "title"],
        ["parties", "items", "properties", "contactPoint", "title"],
        ["parties", "items", "properties", "contactPoint", "properties", "name", "title"],
    ]
    assert headers.get_header("", paths) == "Parties: Organization: Contact point: Name"
